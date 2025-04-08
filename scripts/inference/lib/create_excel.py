import os
import csv
from openpyxl import Workbook
from .excel_headers import TIMING_HEADERS
from .constants import TIMING_FIELDS, EXCEL_SHEETS


def create_csv_file(parallel_mode, file_name="default.csv"):
    """
    Crea o sobrescribe un archivo CSV con las cabeceras por defecto para tiempos.

    :param file_path: Ruta completa del archivo CSV.
    """

    file_path = f"/TFG/excels/{parallel_mode}/aux_files/" + file_name

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Crear o sobrescribir el archivo CSV con las cabeceras
    with open(file_path, mode="w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(TIMING_HEADERS)
    print(f"[CREATE EXCEL] Archivo {file_path} creado o sobrescrito con cabeceras de tiempos.")

    return file_path


def add_row_to_csv(file_path, frame_index, times):
    """
    Añade una fila al archivo CSV con los datos de tiempos.

    :param file_path: Ruta completa del archivo CSV.
    :param frame_index: Índice del frame actual.
    :param times: Diccionario con los tiempos del frame.
    """
    capture_time = times.get(TIMING_FIELDS["CAPTURE"], 0)
    process_time = times.get(TIMING_FIELDS["PROCESSING"], 0)
    preprocess_time = times.get(TIMING_FIELDS["DETECT_FUNCTION"], {}).get(TIMING_FIELDS["PREPROCESS"], 0)
    inference_time = times.get(TIMING_FIELDS["DETECT_FUNCTION"], {}).get(TIMING_FIELDS["INFERENCE"], 0)
    postprocess_time = times.get(TIMING_FIELDS["DETECT_FUNCTION"], {}).get(TIMING_FIELDS["POSTPROCESS"], 0)
    track_time = times.get(TIMING_FIELDS["TRACKING"], 0)
    write_time = times.get(TIMING_FIELDS["WRITING"], 0)
    object_count = times.get(TIMING_FIELDS["OBJECTS_COUNT"], 0)

    total_frame_time = capture_time + process_time + track_time + write_time
    time_per_object = (track_time / object_count * 1000) if object_count > 0 else 0

    row = [
        frame_index,
        format_number(capture_time),
        format_number(process_time),
        format_number(track_time),
        format_number(write_time),
        object_count,
        format_number(total_frame_time * 1000),
        format_number(time_per_object),
        format_number(preprocess_time),
        format_number(inference_time),
        format_number(postprocess_time),
    ]

    # Escribe la fila en el archivo CSV
    with open(file_path, mode="a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(row)

def add_fps_to_csv(file_path, frame_index, fps_value):
    """
    Añade una fila al archivo CSV con el frame y su FPS.
    Si se llama por primera vez, reemplaza los encabezados existentes por los de FPS.

    :param file_path: Ruta completa del archivo CSV.
    :param frame_index: Índice del frame actual.
    :param fps_value: Valor de FPS calculado.
    """
    # Verificar si el archivo tiene los encabezados de tiempos
    rewrite_headers = False
    if os.path.exists(file_path):
        with open(file_path, mode="r", newline="") as f:
            first_row = next(csv.reader(f), [])
            if first_row != [EXCEL_SHEETS["FRAME"], EXCEL_SHEETS["FPS"]]:  # No son los encabezados de FPS
                rewrite_headers = True

    # Si es la primera vez, reescribe los encabezados
    if rewrite_headers:
        with open(file_path, mode="w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([EXCEL_SHEETS["FRAME"], EXCEL_SHEETS["FPS"]])
        # print(f"[CREATE EXCEL] Encabezados de tiempos reemplazados por los de FPS en {file_path}.")

    # Añadir fila con el valor de FPS
    with open(file_path, mode="a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([frame_index, format_number(fps_value)])

def format_number(number):
    """
    Formatea un número al estilo 'es_ES' (con coma como separador decimal).
    """
    return f"{number:.6f}".replace(".", ",")


def create_excel_from_csv(
    times_name, fps_name, hardware_usage_name, parallel_mode, output_name="default.xlsx"
):
    """
    Función que crea un archivo Excel a partir de dos CSVs utilizando openpyxl.

    :param times_name: Nombre del archivo CSV con los tiempos.
    :param fps_name: Nombre del archivo CSV con los FPS.
    :param output_name: Nombre del archivo Excel de salida (incluyendo ruta si es necesario).
    """
    print(f"[CREATE EXCEL] Creando Excel a partir de {times_name} y {fps_name}...")

    # Ruta de los archivos CSV
    aux_file_path = f"/TFG/excels/{parallel_mode}/aux_files/"
    file_path = f"/TFG/excels/{parallel_mode}/"

    # Leer los CSVs
    times_data = []
    with open(aux_file_path + times_name, mode="r") as f:
        reader = csv.reader(f)
        times_data = list(reader)
        
    print(f"[CREATE EXCEL] CSV de tiempos leído correctamente desde {aux_file_path + times_name}.")

    fps_data = []
    with open(aux_file_path + fps_name, mode="r") as f:
        reader = csv.reader(f)
        fps_data = list(reader)
        
    print(f"[CREATE EXCEL] CSV de FPS leído correctamente desde {aux_file_path + fps_name}.")

    hardware_usage_data = []
    with open(aux_file_path + hardware_usage_name, mode="r") as f:
        reader = csv.reader(f)
        hardware_usage_data = list(reader)
        
    print(f"[CREATE EXCEL] CSV de uso de hardware leído correctamente desde {aux_file_path + hardware_usage_name}.")

    # Crear un libro de trabajo y las hojas
    wb = Workbook()

    # Crear la hoja de 'Times'
    times_sheet = wb.active
    times_sheet.title = EXCEL_SHEETS["TIMES"]

    # Escribir los datos de times.csv en la hoja de tiempos
    for row in times_data:
        processed_row = [
            (
                float(cell.replace(",", ".").replace("'", ""))
                if cell.replace(",", "").isdigit()
                else cell
            )
            for cell in row
        ]
        times_sheet.append(processed_row)

    # Crear la hoja de 'FPS'
    fps_sheet = wb.create_sheet(title=EXCEL_SHEETS["FPS"])

    # Escribir los datos de fps.csv en la hoja de FPS
    for row in fps_data:
        processed_row = [
            (
                float(cell.replace(",", ".").replace("'", ""))
                if cell.replace(",", "").isdigit()
                else cell
            )
            for cell in row
        ]
        fps_sheet.append(processed_row)

    # Crear la hoja de 'Hardware Usage'
    hardware_sheet = wb.create_sheet(title=EXCEL_SHEETS["HARDWARE"])
    for row in hardware_usage_data:
        processed_row = [
            (
                float(cell.replace(",", ".").replace("'", ""))
                if cell.replace(",", "").isdigit()
                else cell
            )
            for cell in row
        ]
        hardware_sheet.append(processed_row)

    # Verificar si la ruta de salida existe, si no, crearla
    output_dir = os.path.dirname(output_name)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    import shutil
    print(f"[CREATE EXCEL] Borrando el excel anterior en {file_path + output_name}...")
    try:
        os.remove(file_path + output_name)
    except FileNotFoundError:
        pass
    print(f"[CREATE EXCEL] Excel anterior borrado exitosamente en {file_path + output_name}.")    
    
    
    # Guardar el libro de trabajo como archivo Excel
    print(f"[CREATE EXCEL] Guardando Excel en {file_path + output_name}...")
    wb.save(file_path + output_name)
    print(f"[CREATE EXCEL] Excel guardado exitosamente en {file_path + output_name}.")

    # Opcional: borrar los archivos CSV después de crear el Excel

    shutil.rmtree(f"/TFG/excels/{parallel_mode}/aux_files/")

    print(f"Excel generado exitosamente en {file_path + output_name}.")
