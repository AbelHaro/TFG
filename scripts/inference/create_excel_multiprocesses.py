import os
import openpyxl
import csv

def initialize_excel(file="times.xlsx"):
    """
    Crea un archivo Excel con dos hojas: 'Sheet1' para los tiempos y 'Sheet2' para los FPS.
    """
    output_excels_dir = "/TFG/excels/"
    os.makedirs(output_excels_dir, exist_ok=True)
    output_excel_file = os.path.join(output_excels_dir, file)

    # Crea un libro de trabajo y agrega las hojas
    wb = openpyxl.Workbook()

    # Hoja para los tiempos
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    sheet1.append([
        "Frame", "Captura", "Procesamiento", "Tracking", "Escritura",
        "Cantidad Objetos", "Tiempo Total (ms)", "Tiempo por Objeto Tracking (ms)",
        "Preprocess", "Inference", "Postprocess"
    ])

    # Hoja para los FPS
    sheet2 = wb.create_sheet('Sheet2')
    sheet2.append(["Frame", "FPS"])

    # Guarda el archivo Excel
    wb.save(output_excel_file)
    print(f"Archivo inicializado con cabeceras: {output_excel_file}")
    return output_excel_file


def add_row_to_excel(output_excel_file, frame_index, times):
    """
    Añade una fila a la hoja 'Sheet1' con los datos extraídos de `times`.
    
    :param output_excel_file: Ruta al archivo Excel.
    :param frame_index: Índice del frame actual.
    :param times: Diccionario con los tiempos del frame.
    """
    capture_time = times.get("capture", 0)
    process_time = times.get("processing", 0)
    preprocess_time = times.get("detect_function", {}).get("preprocess", 0)
    inference_time = times.get("detect_function", {}).get("inference", 0)
    postprocess_time = times.get("detect_function", {}).get("postprocess", 0)
    track_time = times.get("tracking", 0)
    write_time = times.get("writing", 0)
    object_count = times.get("objects_count", 0)
    
    total_frame_time = capture_time + process_time + track_time + write_time
    time_per_object = (track_time / object_count * 1000) if object_count > 0 else 0

    row = [
        frame_index,
        format_number(capture_time),
        format_number(process_time),
        format_number(track_time),
        format_number(write_time),
        object_count,
        format_number(total_frame_time * 1000),
        format_number(time_per_object),
        format_number(preprocess_time),
        format_number(inference_time),
        format_number(postprocess_time)
    ]

    # Abre el archivo Excel y selecciona la hoja 'Sheet1'
    wb = openpyxl.load_workbook(output_excel_file)
    sheet1 = wb['Sheet1']
    sheet1.append(row)
    wb.save(output_excel_file)
    print(f"Fila añadida en Sheet1 para el frame {frame_index}")


def add_fps_to_excel(output_excel_file, frame_index, fps_value):
    """
    Añade el valor de FPS en la segunda hoja 'Sheet2' para el frame correspondiente.
    Si ya existe un valor de FPS para ese frame, no lo modifica.
    
    :param output_excel_file: Ruta al archivo Excel.
    :param frame_index: Índice del frame actual.
    :param fps_value: Valor de FPS calculado.
    """
    # Abre el archivo Excel y selecciona la hoja 'Sheet2'
    wb = openpyxl.load_workbook(output_excel_file)
    sheet2 = wb['Sheet2']

    # Verifica si ya existe un valor de FPS para este frame
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, min_col=1, max_col=2):  # Omite la cabecera
        if row[0].value == frame_index:
            if row[1].value is None:  # Si la celda de FPS está vacía, escribe el valor
                row[1].value = format_number(fps_value)
            break
    else:
        # Si no existe, añade una nueva fila
        sheet2.append([frame_index, format_number(fps_value)])
    
    wb.save(output_excel_file)
    print(f"FPS {fps_value} añadido para el frame {frame_index} en Sheet2")


def format_number(number):
    """
    Formatea un número al estilo 'es_ES' (con coma como separador decimal).
    """
    return f"{number:.6f}".replace('.', ',')
